import os
import uuid
import shutil
import base64
import json
import struct
import re
import zipfile
import unicodedata
from typing import List, Optional, Dict, Any
from urllib.parse import quote
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pymysql
import io

# COS（腾讯云对象存储）SDK
from qcloud_cos import CosConfig, CosS3Client
import logging

# 关闭COS SDK的详细日志输出
logging.getLogger('qcloud_cos').setLevel(logging.WARNING)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pypdf
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

app = FastAPI(title="PDF表单填写工具")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 数据库配置 - 支持 MYSQL_ 和 DB_ 两种前缀的环境变量
DB_CONFIG = {
    "host": os.environ.get("MYSQL_HOST", os.environ.get("DB_HOST", "sh-cynosdbmysql-grp-5awkhsnm.sql.tencentcdb.com")),
    "port": int(os.environ.get("MYSQL_PORT", os.environ.get("DB_PORT", 21797))),
    "database": os.environ.get("MYSQL_DATABASE", os.environ.get("DB_NAME", "pdf-generator-prod-8dk636da61e07")),
    "user": os.environ.get("MYSQL_USER", os.environ.get("DB_USER", "admin")),
    "password": os.environ.get("MYSQL_PASSWORD", os.environ.get("DB_PASSWORD", "Kx7#mPqR2@nL9vZw")),
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}

# ==================== COS对象存储配置 ====================
COS_SECRET_ID = os.environ.get("COS_SECRET_ID", "AKIDM3ficMuJXbUAj9NODqJZ8Vr7B5pOmHgH")
COS_SECRET_KEY = os.environ.get("COS_SECRET_KEY", "Z2d6QbdS7KNFuPl9PVTIxRjprjT2OYLx")
COS_REGION = os.environ.get("COS_REGION", "ap-shanghai")
COS_BUCKET = os.environ.get("COS_BUCKET", "pdf-template-1411244947")
COS_PREFIX = os.environ.get("COS_PREFIX", "pdf-templates/")  # COS中的存储前缀（目录）

# 初始化COS客户端
_cos_client = None

def _get_cos_client() -> CosS3Client:
    """获取COS客户端（延迟初始化，单例）"""
    global _cos_client
    if _cos_client is None:
        if not COS_SECRET_ID or not COS_SECRET_KEY or not COS_BUCKET:
            raise RuntimeError("COS配置不完整，请检查环境变量 COS_SECRET_ID / COS_SECRET_KEY / COS_BUCKET")
        config = CosConfig(
            Region=COS_REGION,
            SecretId=COS_SECRET_ID,
            SecretKey=COS_SECRET_KEY,
            Token=None,
            Scheme="https",
        )
        _cos_client = CosS3Client(config)
        print(f"COS客户端已初始化: Bucket={COS_BUCKET}, Region={COS_REGION}, Prefix={COS_PREFIX}")
    return _cos_client


def get_db():
    return pymysql.connect(**DB_CONFIG)


def _save_pdf_file(pdf_bytes: bytes, original_filename: str = "") -> str:
    """将PDF文件上传到COS，返回COS对象Key（相对路径）"""
    ext = os.path.splitext(original_filename)[1] if original_filename else ".pdf"
    if not ext:
        ext = ".pdf"
    file_id = uuid.uuid4().hex
    # 按日期分目录存储
    from datetime import datetime
    date_dir = datetime.now().strftime("%Y%m")
    filename = f"{file_id}{ext}"
    cos_key = f"{COS_PREFIX}{date_dir}/{filename}"

    client = _get_cos_client()
    client.put_object(
        Bucket=COS_BUCKET,
        Body=pdf_bytes,
        Key=cos_key,
        ContentType="application/pdf",
    )
    print(f"PDF已上传到COS: {cos_key} ({len(pdf_bytes)} bytes)")
    return cos_key


def _read_pdf_file(pdf_path: str) -> bytes:
    """从COS读取PDF文件"""
    client = _get_cos_client()
    try:
        response = client.get_object(
            Bucket=COS_BUCKET,
            Key=pdf_path,
        )
        pdf_bytes = response["Body"].get_raw_stream().read()
        return pdf_bytes
    except Exception as e:
        raise FileNotFoundError(f"COS中PDF文件不存在: {pdf_path}, 错误: {e}")


def _delete_pdf_file(pdf_path: str):
    """从COS删除PDF文件"""
    if not pdf_path:
        return
    try:
        client = _get_cos_client()
        client.delete_object(
            Bucket=COS_BUCKET,
            Key=pdf_path,
        )
        print(f"已从COS删除PDF文件: {pdf_path}")
    except Exception as e:
        print(f"从COS删除PDF文件失败: {pdf_path}, 错误: {e}")

def init_db():
    """初始化数据库表结构（首次启动时自动创建）"""
    db = pymysql.connect(
        host=DB_CONFIG["host"],
        port=DB_CONFIG["port"],
        database=DB_CONFIG["database"],
        user=DB_CONFIG["user"],
        password=DB_CONFIG["password"],
        charset=DB_CONFIG["charset"],
    )
    try:
        with db.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS field_definitions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    label VARCHAR(255) NOT NULL,
                    field_type VARCHAR(50) DEFAULT 'text',
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS templates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    description TEXT,
                    pdf_data LONGBLOB,
                    pdf_path VARCHAR(512),
                    pdf_filename VARCHAR(255),
                    page_count INT DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            # 如果表已存在但没有pdf_path字段，则添加
            try:
                cursor.execute("ALTER TABLE templates ADD COLUMN pdf_path VARCHAR(512) AFTER pdf_data")
                print("已添加 pdf_path 字段")
            except Exception:
                pass  # 字段已存在，忽略
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS template_fields (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    template_id INT NOT NULL,
                    field_def_id INT,
                    custom_label VARCHAR(255),
                    page_num INT DEFAULT 1,
                    x DOUBLE DEFAULT 0,
                    y DOUBLE DEFAULT 0,
                    width DOUBLE DEFAULT 0,
                    height DOUBLE DEFAULT 0,
                    font_size INT DEFAULT 12,
                    sort_order INT DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (template_id) REFERENCES templates(id) ON DELETE CASCADE,
                    FOREIGN KEY (field_def_id) REFERENCES field_definitions(id) ON DELETE SET NULL
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            db.commit()
        print("Database tables initialized successfully.")
        
        # 自动迁移：将数据库中残留的LONGBLOB数据迁移到COS
        _migrate_pdf_to_cos(db)
    except Exception as e:
        print(f"Database initialization error: {e}")
    finally:
        db.close()

def _migrate_pdf_to_cos(db):
    """将数据库中的LONGBLOB PDF数据迁移到COS对象存储"""
    if not COS_SECRET_ID or not COS_SECRET_KEY or not COS_BUCKET:
        print("COS未配置，跳过自动迁移。请设置 COS_SECRET_ID / COS_SECRET_KEY / COS_BUCKET 环境变量后手动迁移。")
        return
    try:
        with db.cursor() as cursor:
            # 查找有pdf_data但没有pdf_path的记录
            cursor.execute("SELECT id, pdf_data, pdf_filename FROM templates WHERE pdf_data IS NOT NULL AND (pdf_path IS NULL OR pdf_path = '')")
            rows = cursor.fetchall()
            if not rows:
                return
            print(f"发现 {len(rows)} 个PDF需要迁移到COS...")
            migrated = 0
            for row in rows:
                try:
                    pdf_bytes = bytes(row["pdf_data"]) if not isinstance(row["pdf_data"], bytes) else row["pdf_data"]
                    if len(pdf_bytes) == 0:
                        continue
                    filename = row.get("pdf_filename", "") or f"template_{row['id']}.pdf"
                    cos_key = _save_pdf_file(pdf_bytes, filename)
                    # 更新数据库：设置pdf_path为COS Key，清空pdf_data以释放数据库空间
                    cursor.execute(
                        "UPDATE templates SET pdf_path=%s, pdf_data=NULL WHERE id=%s",
                        (cos_key, row["id"])
                    )
                    migrated += 1
                    print(f"  已迁移模板 #{row['id']}: {filename} -> COS:{cos_key}")
                except Exception as e:
                    print(f"  迁移模板 #{row['id']} 失败: {e}")
            db.commit()
            if migrated > 0:
                print(f"COS迁移完成：成功 {migrated}/{len(rows)} 个")
    except Exception as e:
        print(f"COS迁移过程出错: {e}")

@app.on_event("startup")
def on_startup():
    init_db()

# ==================== Pydantic Models ====================

class FieldDefinitionCreate(BaseModel):
    name: str
    label: str
    field_type: str = "text"
    description: str = ""

class FieldDefinitionUpdate(BaseModel):
    name: Optional[str] = None
    label: Optional[str] = None
    field_type: Optional[str] = None
    description: Optional[str] = None

class TemplateFieldCreate(BaseModel):
    field_def_id: Optional[int] = None
    custom_label: Optional[str] = None
    page_num: int = 1
    x: float
    y: float
    width: float
    height: float
    font_size: int = 12
    sort_order: int = 0

class TemplateFieldUpdate(BaseModel):
    field_def_id: Optional[int] = None
    custom_label: Optional[str] = None
    page_num: Optional[int] = None
    x: Optional[float] = None
    y: Optional[float] = None
    width: Optional[float] = None
    height: Optional[float] = None
    font_size: Optional[int] = None
    sort_order: Optional[int] = None

class TemplateSaveFields(BaseModel):
    fields: List[TemplateFieldCreate]

# ==================== Field Definitions API ====================

@app.get("/api/field-definitions")
def list_field_definitions():
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT * FROM field_definitions ORDER BY id")
            rows = cursor.fetchall()
        return {"data": rows}
    finally:
        db.close()

@app.post("/api/field-definitions")
def create_field_definition(body: FieldDefinitionCreate):
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute(
                "INSERT INTO field_definitions (name, label, field_type, description) VALUES (%s, %s, %s, %s)",
                (body.name, body.label, body.field_type, body.description)
            )
            db.commit()
            new_id = cursor.lastrowid
            cursor.execute("SELECT * FROM field_definitions WHERE id=%s", (new_id,))
            row = cursor.fetchone()
        return {"data": row}
    finally:
        db.close()

@app.put("/api/field-definitions/{fid}")
def update_field_definition(fid: int, body: FieldDefinitionUpdate):
    db = get_db()
    try:
        with db.cursor() as cursor:
            updates = {}
            if body.name is not None: updates["name"] = body.name
            if body.label is not None: updates["label"] = body.label
            if body.field_type is not None: updates["field_type"] = body.field_type
            if body.description is not None: updates["description"] = body.description
            if not updates:
                raise HTTPException(status_code=400, detail="No fields to update")
            set_clause = ", ".join([f"{k}=%s" for k in updates])
            cursor.execute(
                f"UPDATE field_definitions SET {set_clause} WHERE id=%s",
                list(updates.values()) + [fid]
            )
            db.commit()
            cursor.execute("SELECT * FROM field_definitions WHERE id=%s", (fid,))
            row = cursor.fetchone()
        return {"data": row}
    finally:
        db.close()

@app.delete("/api/field-definitions/{fid}")
def delete_field_definition(fid: int):
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("DELETE FROM field_definitions WHERE id=%s", (fid,))
            db.commit()
        return {"success": True}
    finally:
        db.close()

# ==================== Templates API ====================

@app.get("/api/templates")
def list_templates():
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, name, description, pdf_filename, page_count, created_at, updated_at FROM templates ORDER BY created_at DESC"
            )
            rows = cursor.fetchall()
        return {"data": rows}
    finally:
        db.close()

@app.post("/api/templates")
async def create_template(
    name: str = Form(...),
    description: str = Form(""),
    pdf_file: UploadFile = File(...)
):
    pdf_bytes = await pdf_file.read()
    # 保存PDF到文件系统
    pdf_path = _save_pdf_file(pdf_bytes, pdf_file.filename)
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute(
                "INSERT INTO templates (name, description, pdf_path, pdf_filename, page_count) VALUES (%s, %s, %s, %s, %s)",
                (name, description, pdf_path, pdf_file.filename, 1)
            )
            db.commit()
            new_id = cursor.lastrowid
            cursor.execute(
                "SELECT id, name, description, pdf_filename, pdf_path, page_count, created_at FROM templates WHERE id=%s",
                (new_id,)
            )
            row = cursor.fetchone()
        return {"data": row}
    finally:
        db.close()

@app.get("/api/templates/{tid}/parse-fields")
def parse_template_fields(tid: int):
    """解析PDF中的原生表单字段，返回字段位置和类型信息"""
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT pdf_data, pdf_path, pdf_filename, page_count FROM templates WHERE id=%s", (tid,))
            row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Template not found")

        # 优先从文件系统读取PDF
        pdf_bytes = None
        if row.get("pdf_path"):
            try:
                pdf_bytes = _read_pdf_file(row["pdf_path"])
            except FileNotFoundError:
                pass
        # 兼容旧数据
        if pdf_bytes is None and row.get("pdf_data"):
            pdf_bytes = bytes(row["pdf_data"]) if not isinstance(row["pdf_data"], bytes) else row["pdf_data"]
        
        if not pdf_bytes:
            raise HTTPException(status_code=404, detail="PDF数据不存在")
        
        parsed_fields = []

        if HAS_PYPDF:
            try:
                parsed_fields = _parse_with_pypdf(pdf_bytes)
            except Exception as e:
                print(f"pypdf parse error: {e}")

        # 如果pypdf没解析到字段，尝试pdfplumber
        if not parsed_fields and HAS_PDFPLUMBER:
            try:
                parsed_fields = _parse_with_pdfplumber(pdf_bytes)
            except Exception as e:
                print(f"pdfplumber parse error: {e}")

        return {"data": {"fields": parsed_fields, "total": len(parsed_fields)}}
    finally:
        db.close()


def _parse_with_pypdf(pdf_bytes: bytes) -> list:
    """使用pypdf解析PDF表单字段"""
    import pypdf
    fields_result = []
    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    total_pages = len(reader.pages)

    # 解析AcroForm字段
    if reader.trailer.get("/Root") and reader.trailer["/Root"].get("/AcroForm"):
        acro_form = reader.trailer["/Root"]["/AcroForm"]
        if "/Fields" in acro_form:
            fields = acro_form["/Fields"]
            for i, field_ref in enumerate(fields):
                try:
                    field = field_ref.get_object() if hasattr(field_ref, 'get_object') else field_ref
                    field_info = _extract_field_info_pypdf(field, reader, i)
                    if field_info:
                        fields_result.append(field_info)
                except Exception as e:
                    print(f"Field parse error: {e}")
                    continue

    return fields_result


def _extract_field_info_pypdf(field, reader, idx: int) -> dict:
    """从pypdf字段对象提取字段信息"""
    try:
        # 字段名
        name = ""
        if "/T" in field:
            name = str(field["/T"])
        if not name:
            name = f"field_{idx+1}"

        # 字段类型
        ft = str(field.get("/FT", "/Tx"))
        type_map = {"/Tx": "text", "/Btn": "checkbox", "/Ch": "select", "/Sig": "signature"}
        field_type = type_map.get(ft, "text")

        # 如果是按钮类型，判断是否是复选框
        if ft == "/Btn":
            ff = int(field.get("/Ff", 0))
            if ff & (1 << 15):  # Radio button flag
                field_type = "radio"
            else:
                field_type = "checkbox"

        # 获取矩形位置（需要找到widget annotation）
        rect = None
        page_num = 1

        # 直接从字段获取Rect
        if "/Rect" in field:
            r = field["/Rect"]
            rect = [float(r[0]), float(r[1]), float(r[2]), float(r[3])]

        # 从Kids中获取
        if rect is None and "/Kids" in field:
            kids = field["/Kids"]
            for kid_ref in kids:
                try:
                    kid = kid_ref.get_object() if hasattr(kid_ref, 'get_object') else kid_ref
                    if "/Rect" in kid:
                        r = kid["/Rect"]
                        rect = [float(r[0]), float(r[1]), float(r[2]), float(r[3])]
                        break
                except:
                    continue

        if rect is None:
            return None

        # 查找字段所在页码
        for page_idx, page in enumerate(reader.pages):
            if "/Annots" in page:
                annots = page["/Annots"]
                for annot_ref in annots:
                    try:
                        annot = annot_ref.get_object() if hasattr(annot_ref, 'get_object') else annot_ref
                        if "/T" in annot and str(annot["/T"]) == name:
                            page_num = page_idx + 1
                            if "/Rect" in annot:
                                r = annot["/Rect"]
                                rect = [float(r[0]), float(r[1]), float(r[2]), float(r[3])]
                            break
                    except:
                        continue

        # 计算x, y, width, height（PDF坐标系，原点在左下角）
        x1, y1, x2, y2 = rect
        x = min(x1, x2)
        y = min(y1, y2)
        width = abs(x2 - x1)
        height = abs(y2 - y1)

        # 生成友好的标签名
        label = _generate_label(name)

        return {
            "name": name,
            "label": label,
            "field_type": field_type,
            "page_num": page_num,
            "x": round(x, 2),
            "y": round(y, 2),
            "width": round(width, 2),
            "height": round(height, 2),
            "font_size": 12,
            "source": "acroform",
        }
    except Exception as e:
        print(f"Extract field error: {e}")
        return None


def _parse_with_pdfplumber(pdf_bytes: bytes) -> list:
    """使用pdfplumber解析PDF（备用方案）"""
    import pdfplumber
    fields_result = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            # 尝试获取表单字段
            if hasattr(page, 'annots') and page.annots:
                for annot in page.annots:
                    try:
                        if annot.get('subtype') in ('Widget',):
                            x0 = float(annot.get('x0', 0))
                            y0 = float(annot.get('y0', 0))
                            x1 = float(annot.get('x1', 0))
                            y1 = float(annot.get('y1', 0))
                            name = annot.get('field_name', f'field_{len(fields_result)+1}')
                            fields_result.append({
                                "name": name,
                                "label": _generate_label(name),
                                "field_type": "text",
                                "page_num": page_idx + 1,
                                "x": round(min(x0, x1), 2),
                                "y": round(min(y0, y1), 2),
                                "width": round(abs(x1 - x0), 2),
                                "height": round(abs(y1 - y0), 2),
                                "font_size": 12,
                                "source": "pdfplumber",
                            })
                    except Exception as e:
                        print(f"pdfplumber annot error: {e}")
                        continue
    return fields_result


def _generate_label(name: str) -> str:
    """将字段名转换为友好的中文标签"""
    # 常见字段名映射
    label_map = {
        "name": "姓名", "full_name": "全名", "first_name": "名", "last_name": "姓",
        "email": "邮箱", "phone": "电话", "mobile": "手机号",
        "date": "日期", "start_date": "开始日期", "end_date": "结束日期",
        "address": "地址", "city": "城市", "country": "国家",
        "signature": "签名", "sign": "签名",
        "department": "部门", "dept": "部门",
        "position": "职位", "title": "职称",
        "id": "编号", "no": "编号", "number": "编号",
        "reason": "原因", "description": "描述", "remark": "备注", "comment": "备注",
        "amount": "金额", "total": "合计",
        "approve": "审批", "approved": "已审批",
        "checkbox": "复选框", "check": "勾选",
    }
    lower_name = name.lower().strip()
    if lower_name in label_map:
        return label_map[lower_name]
    # 尝试部分匹配
    for key, val in label_map.items():
        if key in lower_name:
            return val
    # 将下划线/驼峰转为空格
    label = re.sub(r'[_\-]', ' ', name)
    label = re.sub(r'([A-Z])', r' \1', label).strip()
    return label if label else name


@app.put("/api/templates/{tid}")
def update_template(tid: int, name: str = Form(None), description: str = Form(None)):
    db = get_db()
    try:
        with db.cursor() as cursor:
            updates = {}
            if name is not None: updates["name"] = name
            if description is not None: updates["description"] = description
            if updates:
                set_clause = ", ".join([f"{k}=%s" for k in updates])
                cursor.execute(
                    f"UPDATE templates SET {set_clause} WHERE id=%s",
                    list(updates.values()) + [tid]
                )
                db.commit()
            cursor.execute(
                "SELECT id, name, description, pdf_filename, page_count, created_at FROM templates WHERE id=%s",
                (tid,)
            )
            row = cursor.fetchone()
        return {"data": row}
    finally:
        db.close()

@app.delete("/api/templates/{tid}")
def delete_template(tid: int):
    db = get_db()
    try:
        with db.cursor() as cursor:
            # 先获取pdf_path以便删除文件
            cursor.execute("SELECT pdf_path FROM templates WHERE id=%s", (tid,))
            row = cursor.fetchone()
            if row and row.get("pdf_path"):
                _delete_pdf_file(row["pdf_path"])
            cursor.execute("DELETE FROM templates WHERE id=%s", (tid,))
            db.commit()
        return {"success": True}
    finally:
        db.close()

@app.get("/api/templates/{tid}/pdf")
def get_template_pdf(tid: int):
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT pdf_data, pdf_path, pdf_filename FROM templates WHERE id=%s", (tid,))
            row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Template not found")
        
        # 优先从文件系统读取
        if row.get("pdf_path"):
            try:
                pdf_bytes = _read_pdf_file(row["pdf_path"])
                b64 = base64.b64encode(pdf_bytes).decode()
                return {"data": {"pdf_base64": b64, "filename": row["pdf_filename"], "storage": "cos"}}
            except FileNotFoundError:
                pass  # 文件不存在，尝试从数据库读取
        
        # 兼容旧数据：从数据库LONGBLOB读取
        pdf_data = row.get("pdf_data")
        if pdf_data:
            pdf_bytes = bytes(pdf_data) if not isinstance(pdf_data, bytes) else pdf_data
            b64 = base64.b64encode(pdf_bytes).decode()
            return {"data": {"pdf_base64": b64, "filename": row["pdf_filename"], "storage": "database"}}
        
        raise HTTPException(status_code=404, detail="PDF数据不存在")
    finally:
        db.close()

@app.put("/api/templates/{tid}/page-count")
def update_page_count(tid: int, page_count: int = Form(...)):
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("UPDATE templates SET page_count=%s WHERE id=%s", (page_count, tid))
            db.commit()
        return {"success": True}
    finally:
        db.close()

# ==================== Template Fields API ====================

@app.get("/api/templates/{tid}/fields")
def get_template_fields(tid: int):
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("""
                SELECT tf.*, fd.name as field_name, fd.label as field_label, fd.field_type, fd.description
                FROM template_fields tf
                LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                WHERE tf.template_id = %s
                ORDER BY tf.page_num, tf.sort_order, tf.id
            """, (tid,))
            rows = cursor.fetchall()
        return {"data": rows}
    finally:
        db.close()

@app.post("/api/templates/{tid}/fields")
def save_template_fields(tid: int, body: TemplateSaveFields):
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("DELETE FROM template_fields WHERE template_id=%s", (tid,))
            for i, f in enumerate(body.fields):
                cursor.execute("""
                    INSERT INTO template_fields
                    (template_id, field_def_id, custom_label, page_num, x, y, width, height, font_size, sort_order)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    tid, f.field_def_id, f.custom_label,
                    f.page_num, f.x, f.y, f.width, f.height,
                    f.font_size, i
                ))
            db.commit()
            cursor.execute("""
                SELECT tf.*, fd.name as field_name, fd.label as field_label, fd.field_type
                FROM template_fields tf
                LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                WHERE tf.template_id = %s
                ORDER BY tf.page_num, tf.sort_order
            """, (tid,))
            rows = cursor.fetchall()
        return {"data": rows}
    finally:
        db.close()

# ==================== Field Sorting Helpers ====================

# 简易拼音首字母表（覆盖GB2312常用汉字区间）
_PINYIN_TABLE = [
    ('\u0041', 0xB0A1), ('\u0042', 0xB0C5), ('\u0043', 0xB2C1), ('\u0044', 0xB4EE),
    ('\u0045', 0xB6EA), ('\u0046', 0xB7A2), ('\u0047', 0xB8C1), ('\u0048', 0xB9FE),
    ('\u004A', 0xBBF7), ('\u004B', 0xBFA6), ('\u004C', 0xC0AC), ('\u004D', 0xC2E8),
    ('\u004E', 0xC4C3), ('\u004F', 0xC5B6), ('\u0050', 0xC5BE), ('\u0051', 0xC6DA),
    ('\u0052', 0xC8BB), ('\u0053', 0xC8F6), ('\u0054', 0xCBFA), ('\u0057', 0xCDDA),
    ('\u0058', 0xCEF4), ('\u0059', 0xD1B9), ('\u005A', 0xD4D1),
]

def _char_pinyin_key(ch):
    """返回单个字符的拼音排序键：中文返回拼音首字母，其他字符返回小写形式"""
    if '\u4e00' <= ch <= '\u9fff':
        try:
            gb_bytes = ch.encode('gb2312')
            code = gb_bytes[0] * 256 + gb_bytes[1]
            for i in range(len(_PINYIN_TABLE) - 1, -1, -1):
                if code >= _PINYIN_TABLE[i][1]:
                    return _PINYIN_TABLE[i][0].lower()
            return 'a'
        except (UnicodeEncodeError, IndexError):
            return ch.lower()
    return ch.lower()

def _pinyin_sort_key(text):
    """生成文本的拼音排序键"""
    return ''.join(_char_pinyin_key(ch) for ch in text)

def _sort_fields_for_excel(fields, field_labels):
    """
    智能排序字段用于Excel模板显示：
    1. 固定字段置顶：保单持有人姓名 排第一，保单持有人证件号码 排第二
    2. 按编号分组：所有 #1 的放一起，所有 #2 的放一起，依此类推
    3. 每组内按拼音排序
    4. 没有编号的字段排在最后，按拼音排序
    
    返回排序后的 (fields, field_labels) 元组
    """
    if not fields:
        return fields, field_labels
    
    # 固定置顶字段（按顺序）
    PIN_TOP = ['保单持有人姓名', '保单持有人证件号码']
    
    pinned = []       # 置顶字段
    numbered = {}     # 编号 -> [(index, field, label, base_label)]
    unnumbered = []   # 无编号字段
    
    for i, (field, label) in enumerate(zip(fields, field_labels)):
        # 提取基础标签（去掉 #N 后缀）
        match = re.match(r'^(.+?)\s*#(\d+)$', label)
        base_label = match.group(1).strip() if match else label
        
        # 检查是否是置顶字段
        if base_label in PIN_TOP:
            pinned.append((i, field, label, base_label))
        elif match:
            num = int(match.group(2))
            if num not in numbered:
                numbered[num] = []
            numbered[num].append((i, field, label, base_label))
        else:
            unnumbered.append((i, field, label))
    
    # 组装结果
    sorted_fields = []
    sorted_labels = []
    
    # 1. 置顶字段按预定义顺序
    pinned.sort(key=lambda x: PIN_TOP.index(x[3]) if x[3] in PIN_TOP else 999)
    for _, field, label, _ in pinned:
        sorted_fields.append(field)
        sorted_labels.append(label)
    
    # 2. 按编号分组：#1 组、#2 组...，每组内按拼音排序
    for num in sorted(numbered.keys()):
        group = numbered[num]
        group.sort(key=lambda x: _pinyin_sort_key(x[3]))  # 按基础标签拼音排
        for _, field, label, _ in group:
            sorted_fields.append(field)
            sorted_labels.append(label)
    
    # 3. 无编号字段按拼音排序
    unnumbered.sort(key=lambda x: _pinyin_sort_key(x[2]))
    for _, field, label in unnumbered:
        sorted_fields.append(field)
        sorted_labels.append(label)
    
    return sorted_fields, sorted_labels

# ==================== Excel Batch API ====================

@app.get("/api/templates/{tid}/excel-template")
def download_excel_template(tid: int):
    """生成并下载Excel填写模板（含下拉选项和单元格启用/禁用）"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, name FROM templates WHERE id=%s", (tid,))
            tmpl = cursor.fetchone()
            if not tmpl:
                raise HTTPException(status_code=404, detail="Template not found")
            cursor.execute("""
                SELECT tf.*, fd.name as field_name, fd.label as field_label, fd.field_type, fd.description
                FROM template_fields tf
                LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                WHERE tf.template_id = %s
                ORDER BY tf.page_num, tf.sort_order, tf.id
            """, (tid,))
            fields = cursor.fetchall()
    finally:
        db.close()

    if not fields:
        raise HTTPException(status_code=400, detail="此模板没有配置字段，请先在管理后台配置字段")

    # 为重复字段添加编号
    label_counts = {}
    for field in fields:
        label = field.get('custom_label') or field.get('field_label') or field.get('field_name') or f'字段'
        label_counts[label] = label_counts.get(label, 0) + 1

    # 生成带编号的标签列表
    label_seen = {}
    field_labels = []
    for field in fields:
        label = field.get('custom_label') or field.get('field_label') or field.get('field_name') or f'字段'
        if label_counts[label] > 1:
            label_seen[label] = label_seen.get(label, 0) + 1
            display_label = f"{label} #{label_seen[label]}"
        else:
            display_label = label
        field_labels.append(display_label)

    # 智能排序：有编号的分组显示，没编号的按拼音排序
    fields, field_labels = _sort_fields_for_excel(fields, field_labels)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "填写数据"

    # 样式定义
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    required_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    disabled_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    enabled_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # 第一行：序号列 + 字段列
    ws.column_dimensions['A'].width = 8
    ws.cell(row=1, column=1, value="序号").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=1).border = thin_border

    # 第二行：字段类型说明
    ws.cell(row=2, column=1, value="类型").fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    ws.cell(row=2, column=1).font = Font(color="5B21B6", bold=True, size=9)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).border = thin_border

    # 第三行：字段说明
    ws.cell(row=3, column=1, value="说明").fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    ws.cell(row=3, column=1).font = Font(color="5B21B6", bold=True, size=9)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1).border = thin_border

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 字段类型映射
    type_labels = {
        "text": "文本", "textarea": "多行文本", "date": "日期",
        "number": "数字", "checkbox": "复选框(是/否)", "select": "下拉选择"
    }

    # 下拉选项sheet（用于存放下拉数据）
    ws_options = wb.create_sheet("选项数据")
    ws_options.sheet_state = 'hidden'  # 隐藏选项数据sheet

    option_col = 1  # 选项数据sheet的列索引

    for col_idx, field in enumerate(fields, start=2):
        label = field_labels[col_idx - 2]
        field_type = field.get('field_type') or 'text'
        col_letter = get_column_letter(col_idx)

        # 设置列宽
        ws.column_dimensions[col_letter].width = max(15, len(label) * 2 + 4)

        # 第1行：字段标签（表头）
        header_cell = ws.cell(row=1, column=col_idx, value=label)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_align
        header_cell.border = thin_border

        # 第2行：字段类型
        type_cell = ws.cell(row=2, column=col_idx, value=type_labels.get(field_type, field_type))
        type_cell.fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
        type_cell.font = Font(color="5B21B6", size=9)
        type_cell.alignment = Alignment(horizontal="center", vertical="center")
        type_cell.border = thin_border

        # 第3行：填写说明
        hint_map = {
            "text": "请输入文本",
            "textarea": "请输入多行文本",
            "date": "格式：YYYY-MM-DD",
            "number": "请输入数字",
            "checkbox": "请选择：是 或 否",
            "select": "请从下拉选择"
        }
        hint_cell = ws.cell(row=3, column=col_idx, value=hint_map.get(field_type, "请填写"))
        hint_cell.fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")
        hint_cell.font = Font(color="6B7280", size=9, italic=True)
        hint_cell.alignment = Alignment(horizontal="center", vertical="center")
        hint_cell.border = thin_border

        # 数据行（第4行起，预留50行）
        for data_row in range(4, 54):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = thin_border

            if field_type == 'checkbox':
                # 复选框：下拉选择 是/否
                cell.fill = enabled_fill
                dv = DataValidation(
                    type="list",
                    formula1='"是,否"',
                    allow_blank=True,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="输入错误",
                    error="请选择 是 或 否"
                )
                dv.sqref = f"{col_letter}{data_row}"
                ws.add_data_validation(dv)

            elif field_type == 'date':
                # 日期格式
                cell.fill = enabled_fill
                cell.number_format = 'YYYY-MM-DD'
                dv = DataValidation(
                    type="date",
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="日期格式错误",
                    error="请输入有效日期"
                )
                dv.sqref = f"{col_letter}{data_row}"
                ws.add_data_validation(dv)

            elif field_type == 'number':
                # 数字格式
                cell.fill = enabled_fill
                cell.number_format = '0.##'

            elif field_type == 'select':
                # 下拉选择（从选项数据sheet读取）
                cell.fill = enabled_fill
                # 将选项写入隐藏sheet
                options = ["选项A", "选项B", "选项C"]  # 默认选项，可扩展
                for opt_row, opt_val in enumerate(options, start=1):
                    ws_options.cell(row=opt_row, column=option_col, value=opt_val)
                opt_range = f"选项数据!${get_column_letter(option_col)}$1:${get_column_letter(option_col)}${len(options)}"
                dv = DataValidation(
                    type="list",
                    formula1=opt_range,
                    allow_blank=True,
                    showDropDown=False
                )
                dv.sqref = f"{col_letter}{data_row}"
                ws.add_data_validation(dv)
                option_col += 1

            else:
                # 普通文本
                cell.fill = enabled_fill

    # 序号列数据行
    for data_row in range(4, 54):
        cell = ws.cell(row=data_row, column=1, value=data_row - 3)
        cell.fill = disabled_fill
        cell.font = Font(color="9CA3AF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 冻结前3行和第1列
    ws.freeze_panes = "B4"

    # 添加说明sheet
    ws_info = wb.create_sheet("使用说明", 0)
    ws_info.column_dimensions['A'].width = 60
    ws_info.row_dimensions[1].height = 40
    info_data = [
        ("PDF批量填写模板使用说明", "FFFFFF", "7C3AED", True, 16),
        ("", "FFFFFF", "FFFFFF", False, 11),
        (f"模板名称：{tmpl['name']}", "1F2937", "F9FAFB", True, 12),
        (f"字段数量：{len(fields)} 个", "1F2937", "F9FAFB", False, 11),
        ("", "FFFFFF", "FFFFFF", False, 11),
        ("【填写步骤】", "7C3AED", "EDE9FE", True, 12),
        ("1. 切换到「填写数据」工作表", "374151", "FFFFFF", False, 11),
        ("2. 从第4行开始逐行填写数据（每行对应一份PDF）", "374151", "FFFFFF", False, 11),
        ("3. 带下拉箭头的单元格请从下拉菜单选择", "374151", "FFFFFF", False, 11),
        ("4. 日期格式请使用 YYYY-MM-DD（如：2024-01-15）", "374151", "FFFFFF", False, 11),
        ("5. 复选框字段请填写「是」或「否」", "374151", "FFFFFF", False, 11),
        ("6. 填写完成后保存Excel文件", "374151", "FFFFFF", False, 11),
        ("7. 在系统中上传此Excel文件，系统将自动批量生成PDF", "374151", "FFFFFF", False, 11),
        ("", "FFFFFF", "FFFFFF", False, 11),
        ("【注意事项】", "DC2626", "FEF2F2", True, 12),
        ("• 请勿修改第1-3行的表头内容", "374151", "FFFFFF", False, 11),
        ("• 请勿删除或移动工作表", "374151", "FFFFFF", False, 11),
        ("• 序号列（A列）无需填写，系统自动生成", "374151", "FFFFFF", False, 11),
        ("• 灰色背景单元格为禁用状态，请勿填写", "374151", "FFFFFF", False, 11),
    ]
    for row_idx, (text, font_color, bg_color, bold, size) in enumerate(info_data, start=1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        cell.font = Font(color=font_color, bold=bold, size=size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws_info.row_dimensions[row_idx].height = 22

    # 将说明sheet设为默认激活
    wb.active = wb["填写数据"]

    # 输出Excel文件
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{tmpl['name']}_批量填写模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/unified-excel-template")
def download_unified_excel_template():
    """生成统一Excel模板：包含所有模板字段的并集，第一列为模板选择，不适用字段灰色"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    db = get_db()
    try:
        with db.cursor() as cursor:
            # 获取所有模板
            cursor.execute("SELECT id, name FROM templates ORDER BY id")
            templates = cursor.fetchall()
            if not templates:
                raise HTTPException(status_code=400, detail="暂无可用模板")

            # 获取所有模板的字段
            template_fields_map = {}  # { template_id: [fields] }
            for tmpl in templates:
                cursor.execute("""
                    SELECT tf.*, fd.name as field_name, fd.label as field_label, fd.field_type, fd.description
                    FROM template_fields tf
                    LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                    WHERE tf.template_id = %s
                    ORDER BY tf.page_num, tf.sort_order, tf.id
                """, (tmpl['id'],))
                fields = cursor.fetchall()
                template_fields_map[tmpl['id']] = fields
    finally:
        db.close()

    # 去掉没有字段的模板
    templates = [t for t in templates if template_fields_map.get(t['id'])]
    if not templates:
        raise HTTPException(status_code=400, detail="所有模板都尚未配置字段，请先在管理后台配置")

    # 构建字段并集：按 (field_label, field_type) 去重，保留顺序
    # 同时记录每个统一字段属于哪些模板
    unified_fields = []  # [{ label, field_type, template_ids: set, description }]
    seen_labels = {}  # label -> index in unified_fields

    for tmpl in templates:
        fields = template_fields_map[tmpl['id']]
        # 为重复字段添加编号
        label_counts = {}
        for field in fields:
            lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or '字段'
            label_counts[lbl] = label_counts.get(lbl, 0) + 1
        label_seen = {}
        for field in fields:
            lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or '字段'
            if label_counts[lbl] > 1:
                label_seen[lbl] = label_seen.get(lbl, 0) + 1
                display_label = f"{lbl} #{label_seen[lbl]}"
            else:
                display_label = lbl

            field_type = field.get('field_type') or 'text'
            # 用 display_label 作为唯一键
            unified_key = display_label

            if unified_key not in seen_labels:
                seen_labels[unified_key] = len(unified_fields)
                unified_fields.append({
                    'label': display_label,
                    'field_type': field_type,
                    'template_ids': {tmpl['id']},
                    'description': field.get('description', ''),
                })
            else:
                unified_fields[seen_labels[unified_key]]['template_ids'].add(tmpl['id'])

    if not unified_fields:
        raise HTTPException(status_code=400, detail="没有可用的字段")

    # 智能排序统一字段：有编号的分组显示，没编号的按拼音排序
    unified_labels = [uf['label'] for uf in unified_fields]
    unified_fields, unified_labels = _sort_fields_for_excel(unified_fields, unified_labels)
    # 更新 seen_labels 索引（排序后位置变了）
    seen_labels = {uf['label']: i for i, uf in enumerate(unified_fields)}

    # ==================== 生成Excel ====================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "填写数据"

    # 样式定义
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    disabled_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    disabled_font = Font(color="9CA3AF", size=10)
    enabled_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    template_col_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    template_col_header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    type_labels = {
        "text": "文本", "textarea": "多行文本", "date": "日期",
        "number": "数字", "checkbox": "复选框(是/否)", "select": "下拉选择"
    }

    # A列：序号
    ws.column_dimensions['A'].width = 8
    ws.cell(row=1, column=1, value="序号").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=1).border = thin_border

    ws.cell(row=2, column=1, value="类型").fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    ws.cell(row=2, column=1).font = Font(color="5B21B6", bold=True, size=9)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).border = thin_border

    ws.cell(row=3, column=1, value="说明").fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    ws.cell(row=3, column=1).font = Font(color="5B21B6", bold=True, size=9)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1).border = thin_border

    # B列：选择模板（下拉）
    ws.column_dimensions['B'].width = 24
    ws.cell(row=1, column=2, value="选择模板 ★").fill = template_col_header_fill
    ws.cell(row=1, column=2).font = Font(color="FFFFFF", bold=True, size=11)
    ws.cell(row=1, column=2).alignment = header_align
    ws.cell(row=1, column=2).border = thin_border

    ws.cell(row=2, column=2, value="必选").fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws.cell(row=2, column=2).font = Font(color="2563EB", bold=True, size=9)
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=2).border = thin_border

    ws.cell(row=3, column=2, value="请从下拉选择模板").fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    ws.cell(row=3, column=2).font = Font(color="6B7280", size=9, italic=True)
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=2).border = thin_border

    # 隐藏sheet存模板选项
    ws_options = wb.create_sheet("选项数据")
    ws_options.sheet_state = 'hidden'

    # 写入模板名称列表到隐藏sheet
    template_names = [t['name'] for t in templates]
    for i, name in enumerate(template_names, start=1):
        ws_options.cell(row=i, column=1, value=name)

    # 构建模板名到ID的映射（写入隐藏sheet第二列供参考）
    for i, t in enumerate(templates, start=1):
        ws_options.cell(row=i, column=2, value=t['id'])

    # 模板下拉验证
    tmpl_opt_range = f"选项数据!$A$1:$A${len(template_names)}"
    option_col_idx = 3  # 选项数据sheet的列索引，从3开始（1,2已被模板占用）

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 构建模板名 -> 模板对象映射
    tmpl_name_to_id = {t['name']: t['id'] for t in templates}
    # 构建模板ID -> 所属统一字段标签集合映射
    tmpl_id_to_labels = {}
    for uf in unified_fields:
        for tid in uf['template_ids']:
            if tid not in tmpl_id_to_labels:
                tmpl_id_to_labels[tid] = set()
            tmpl_id_to_labels[tid].add(uf['label'])

    # C列起：字段列
    for col_offset, uf in enumerate(unified_fields):
        col_idx = col_offset + 3  # 从第3列开始（A=1序号, B=2模板选择）
        col_letter = get_column_letter(col_idx)
        label = uf['label']
        field_type = uf['field_type']

        ws.column_dimensions[col_letter].width = max(15, len(label) * 2 + 4)

        # 第1行：表头
        header_cell = ws.cell(row=1, column=col_idx, value=label)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_align
        header_cell.border = thin_border

        # 第2行：类型 + 适用模板数
        applicable_count = len(uf['template_ids'])
        type_text = type_labels.get(field_type, field_type)
        type_cell = ws.cell(row=2, column=col_idx, value=f"{type_text}（{applicable_count}个模板）")
        type_cell.fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
        type_cell.font = Font(color="5B21B6", size=9)
        type_cell.alignment = Alignment(horizontal="center", vertical="center")
        type_cell.border = thin_border

        # 第3行：适用模板名
        applicable_names = [t['name'] for t in templates if t['id'] in uf['template_ids']]
        hint_text = "适用：" + "、".join(applicable_names)
        hint_cell = ws.cell(row=3, column=col_idx, value=hint_text)
        hint_cell.fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")
        hint_cell.font = Font(color="6B7280", size=8, italic=True)
        hint_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hint_cell.border = thin_border

    # 数据行（第4行起，预留50行）
    for data_row in range(4, 54):
        # 序号列
        cell_seq = ws.cell(row=data_row, column=1, value=data_row - 3)
        cell_seq.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        cell_seq.font = Font(color="9CA3AF", size=10)
        cell_seq.alignment = Alignment(horizontal="center", vertical="center")
        cell_seq.border = thin_border

        # 模板选择列（B列）- 下拉
        tmpl_cell = ws.cell(row=data_row, column=2)
        tmpl_cell.fill = template_col_fill
        tmpl_cell.border = thin_border
        tmpl_cell.font = Font(color="1E40AF", size=10, bold=True)
        dv_tmpl = DataValidation(
            type="list",
            formula1=tmpl_opt_range,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="选择错误",
            error="请从下拉列表中选择模板"
        )
        dv_tmpl.sqref = f"B{data_row}"
        ws.add_data_validation(dv_tmpl)

        # 字段列：默认全部灰色（不适用），等用户选模板后需手动适配
        # 注：Excel无法真正动态变色，但我们用条件格式+注释提示
        for col_offset, uf in enumerate(unified_fields):
            col_idx = col_offset + 3
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = thin_border
            # 默认灰色样式（表示未选模板时不可填写）
            cell.fill = disabled_fill
            cell.font = disabled_font

            field_type = uf['field_type']

            # 添加数据验证（复选框/日期等）
            col_letter = get_column_letter(col_idx)
            if field_type == 'checkbox':
                dv = DataValidation(
                    type="list", formula1='"是,否"', allow_blank=True,
                    showDropDown=False, showErrorMessage=True,
                    errorTitle="输入错误", error="请选择 是 或 否"
                )
                dv.sqref = f"{col_letter}{data_row}"
                ws.add_data_validation(dv)
            elif field_type == 'date':
                cell.number_format = 'YYYY-MM-DD'
            elif field_type == 'number':
                cell.number_format = '0.##'

    # 使用条件格式：当B列选择了某模板后，适用字段变白色，不适用字段保持灰色
    # Excel 条件格式无法直接引用模板-字段映射，所以我们用一个隐藏的映射sheet
    # 创建映射sheet
    ws_map = wb.create_sheet("字段映射")
    ws_map.sheet_state = 'hidden'

    # 在映射sheet中创建矩阵：行=模板，列=统一字段，值=1(适用)/0(不适用)
    # 第1行：字段标签
    for col_offset, uf in enumerate(unified_fields):
        ws_map.cell(row=1, column=col_offset + 2, value=uf['label'])
    # 第1列：模板名称
    for row_offset, tmpl in enumerate(templates):
        ws_map.cell(row=row_offset + 2, column=1, value=tmpl['name'])
        for col_offset, uf in enumerate(unified_fields):
            is_applicable = 1 if tmpl['id'] in uf['template_ids'] else 0
            ws_map.cell(row=row_offset + 2, column=col_offset + 2, value=is_applicable)

    # 为每个字段列添加条件格式：
    # 如果当前行的B列（选择模板）匹配某个模板，且该字段适用于该模板，则背景白色
    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    for col_offset, uf in enumerate(unified_fields):
        col_idx = col_offset + 3
        col_letter = get_column_letter(col_idx)
        map_col_letter = get_column_letter(col_offset + 2)

        # 条件格式：白色（适用）- 当B列选了模板，且对应映射值为1
        # VLOOKUP公式：查找B列的模板名在映射sheet中的行，返回对应字段列的值
        # 如果值=1则适用（白色），否则不适用（灰色）
        formula_applicable = f'AND($B4<>"",VLOOKUP($B4,字段映射!$A:${get_column_letter(len(unified_fields) + 1)},{ col_offset + 2},FALSE)=1)'
        formula_not_applicable = f'AND($B4<>"",VLOOKUP($B4,字段映射!$A:${get_column_letter(len(unified_fields) + 1)},{ col_offset + 2},FALSE)=0)'

        # 适用字段：白色背景
        rule_white = FormulaRule(
            formula=[formula_applicable],
            fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            font=Font(color="000000", size=10),
        )
        ws.conditional_formatting.add(f"{col_letter}4:{col_letter}53", rule_white)

        # 不适用字段：灰色背景 + 灰色字体
        rule_gray = FormulaRule(
            formula=[formula_not_applicable],
            fill=PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
            font=Font(color="9CA3AF", size=10),
        )
        ws.conditional_formatting.add(f"{col_letter}4:{col_letter}53", rule_gray)

    # 冻结前3行和前2列
    ws.freeze_panes = "C4"

    # 使用说明sheet
    ws_info = wb.create_sheet("使用说明", 0)
    ws_info.column_dimensions['A'].width = 70
    ws_info.row_dimensions[1].height = 40

    tmpl_list_text = "、".join([f"「{t['name']}」" for t in templates])
    info_data = [
        ("PDF批量填写模板使用说明（统一模板）", "FFFFFF", "7C3AED", True, 16),
        ("", "FFFFFF", "FFFFFF", False, 11),
        (f"包含模板：{tmpl_list_text}", "1F2937", "F9FAFB", True, 11),
        (f"统一字段数量：{len(unified_fields)} 个（所有模板字段的并集）", "1F2937", "F9FAFB", False, 11),
        ("", "FFFFFF", "FFFFFF", False, 11),
        ("【填写步骤】", "7C3AED", "EDE9FE", True, 12),
        ("1. 切换到「填写数据」工作表", "374151", "FFFFFF", False, 11),
        ("2. 每行的 B 列「选择模板」处，从下拉菜单选择要使用的PDF模板", "374151", "FFFFFF", False, 11),
        ("3. 选择模板后，适用字段列会自动变为白色背景，灰色背景的列表示不适用", "374151", "FFFFFF", False, 11),
        ("4. 仅需填写白色背景的单元格，灰色单元格请留空", "374151", "FFFFFF", False, 11),
        ("5. 带下拉箭头的单元格请从下拉菜单选择", "374151", "FFFFFF", False, 11),
        ("6. 日期格式请使用 YYYY-MM-DD（如：2024-01-15）", "374151", "FFFFFF", False, 11),
        ("7. 复选框字段请填写「是」或「否」", "374151", "FFFFFF", False, 11),
        ("8. 不同行可以选择不同的模板，系统将自动生成对应的PDF", "374151", "FFFFFF", False, 11),
        ("9. 填写完成后保存Excel文件并上传", "374151", "FFFFFF", False, 11),
        ("", "FFFFFF", "FFFFFF", False, 11),
        ("【颜色说明】", "2563EB", "DBEAFE", True, 12),
        ("• 蓝色底色（B列）= 模板选择列（必填）", "374151", "FFFFFF", False, 11),
        ("• 白色底色 = 该字段适用于所选模板，可以填写", "374151", "FFFFFF", False, 11),
        ("• 灰色底色 = 该字段不适用于所选模板，请勿填写", "374151", "FFFFFF", False, 11),
        ("", "FFFFFF", "FFFFFF", False, 11),
        ("【注意事项】", "DC2626", "FEF2F2", True, 12),
        ("• 请勿修改第1-3行的表头内容", "374151", "FFFFFF", False, 11),
        ("• 请勿删除或移动工作表", "374151", "FFFFFF", False, 11),
        ("• 序号列（A列）无需填写", "374151", "FFFFFF", False, 11),
        ("• 如果灰色单元格中填写了内容，系统将自动忽略该内容", "374151", "FFFFFF", False, 11),
    ]
    for row_idx, (text, font_color, bg_color, bold, size) in enumerate(info_data, start=1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        cell.font = Font(color=font_color, bold=bold, size=size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws_info.row_dimensions[row_idx].height = 22

    wb.active = wb["填写数据"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "统一批量填写模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.post("/api/unified-batch-upload")
async def unified_batch_upload(excel_file: UploadFile = File(...)):
    """统一批量上传：从Excel中读取模板选择和数据，按模板分组生成PDF，返回ZIP"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # 读取Excel
    excel_bytes = await excel_file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["填写数据"]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel文件格式错误：{str(e)}")

    # 解析表头（第1行，从C列开始为字段）
    headers = {}
    for col_idx in range(3, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val:
            headers[col_idx] = str(header_val).strip()

    # 读取数据行：每行需要有B列的模板选择
    data_rows = []  # [{ template_name, fields: { label: value } }]
    for row_idx in range(4, ws.max_row + 1):
        template_name = ws.cell(row=row_idx, column=2).value
        if not template_name or not str(template_name).strip():
            continue
        template_name = str(template_name).strip()

        row_fields = {}
        has_data = False
        for col_idx, label in headers.items():
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and str(cell_val).strip():
                row_fields[label] = str(cell_val).strip()
                has_data = True
            else:
                row_fields[label] = ''

        if has_data:
            data_rows.append({
                'template_name': template_name,
                'fields': row_fields,
                'row_num': row_idx,
            })

    if not data_rows:
        raise HTTPException(status_code=400, detail="Excel中没有填写数据，请先选择模板并填写内容（从第4行开始）")

    if len(data_rows) > 100:
        raise HTTPException(status_code=400, detail="单次批量上传最多支持100行数据")

    # 获取所有涉及的模板名称
    needed_template_names = set(r['template_name'] for r in data_rows)

    db = get_db()
    try:
        with db.cursor() as cursor:
            # 查找所有模板
            format_strings = ','.join(['%s'] * len(needed_template_names))
            cursor.execute(
                f"SELECT id, name, pdf_data, pdf_path FROM templates WHERE name IN ({format_strings})",
                list(needed_template_names)
            )
            templates_db = cursor.fetchall()
            tmpl_name_map = {t['name']: t for t in templates_db}

            # 获取每个模板的字段
            template_fields_map = {}
            for tmpl in templates_db:
                cursor.execute("""
                    SELECT tf.*, fd.name as field_name, fd.label as field_label, fd.field_type
                    FROM template_fields tf
                    LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                    WHERE tf.template_id = %s
                    ORDER BY tf.page_num, tf.sort_order, tf.id
                """, (tmpl['id'],))
                template_fields_map[tmpl['id']] = cursor.fetchall()
    finally:
        db.close()

    # 检查是否所有模板都存在
    missing = needed_template_names - set(tmpl_name_map.keys())
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"以下模板名称无法匹配：{'、'.join(missing)}，请检查Excel中B列的模板名称"
        )

    # 按模板分组生成PDF
    zip_buffer = io.BytesIO()
    generated_count = 0
    errors = []

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for row_data in data_rows:
            tmpl = tmpl_name_map.get(row_data['template_name'])
            if not tmpl:
                errors.append(f"第{row_data['row_num']}行：模板「{row_data['template_name']}」不存在")
                continue

            fields = template_fields_map.get(tmpl['id'], [])
            if not fields:
                errors.append(f"第{row_data['row_num']}行：模板「{tmpl['name']}」没有配置字段")
                continue

            pdf_bytes_orig = None
            if tmpl.get("pdf_path"):
                try:
                    pdf_bytes_orig = _read_pdf_file(tmpl["pdf_path"])
                except FileNotFoundError:
                    pass
            if pdf_bytes_orig is None and tmpl.get("pdf_data"):
                pdf_bytes_orig = bytes(tmpl["pdf_data"]) if not isinstance(tmpl["pdf_data"], bytes) else tmpl["pdf_data"]
            if not pdf_bytes_orig:
                errors.append(f"第{row_data['row_num']}行：模板「{tmpl['name']}」的PDF文件不存在")
                continue

            # 构建字段值映射（将统一标签映射到模板字段）
            field_label_counts = {}
            for field in fields:
                lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
                field_label_counts[lbl] = field_label_counts.get(lbl, 0) + 1

            field_label_seen = {}
            field_value_map = {}
            for field in fields:
                lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
                if field_label_counts[lbl] > 1:
                    field_label_seen[lbl] = field_label_seen.get(lbl, 0) + 1
                    display_lbl = f"{lbl} #{field_label_seen[lbl]}"
                else:
                    display_lbl = lbl
                # 从Excel数据中取值
                field_value_map[display_lbl] = row_data['fields'].get(display_lbl, '')

            try:
                filled_pdf = _fill_pdf_with_data(pdf_bytes_orig, fields, field_value_map)
                holder_name = _extract_holder_name(field_value_map)
                seq_num = row_data['row_num'] - 3
                filename = f"{holder_name}_{tmpl['name']}_{seq_num:03d}.pdf" if holder_name else f"{tmpl['name']}_{seq_num:03d}.pdf"
                zf.writestr(filename, filled_pdf)
                generated_count += 1
            except Exception as e:
                errors.append(f"第{row_data['row_num']}行：{str(e)}")

    zip_buffer.seek(0)

    if generated_count == 0:
        raise HTTPException(status_code=500, detail=f"所有PDF生成失败：{'; '.join(errors[:3])}")

    zip_filename = f"批量PDF_{generated_count}份.zip"
    encoded_zip_filename = quote(zip_filename)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_zip_filename}",
            "X-Generated-Count": str(generated_count),
            "X-Error-Count": str(len(errors)),
            "X-Errors": json.dumps(errors[:5], ensure_ascii=False) if errors else "[]",
        }
    )


@app.post("/api/templates/{tid}/batch-upload")
async def batch_upload_excel(tid: int, excel_file: UploadFile = File(...)):
    """上传Excel文件，解析数据并批量生成PDF，返回ZIP包"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, name, pdf_data, pdf_path FROM templates WHERE id=%s", (tid,))
            tmpl = cursor.fetchone()
            if not tmpl:
                raise HTTPException(status_code=404, detail="Template not found")
            cursor.execute("""
                SELECT tf.*, fd.name as field_name, fd.label as field_label, fd.field_type
                FROM template_fields tf
                LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                WHERE tf.template_id = %s
                ORDER BY tf.page_num, tf.sort_order, tf.id
            """, (tid,))
            fields = cursor.fetchall()
    finally:
        db.close()

    if not fields:
        raise HTTPException(status_code=400, detail="此模板没有配置字段")

    # 读取Excel
    excel_bytes = await excel_file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["填写数据"]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel文件格式错误：{str(e)}")

    # 解析表头（第1行）
    headers = {}
    for col_idx in range(2, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val:
            headers[str(header_val).strip()] = col_idx

    # 构建字段映射（支持带编号的标签）
    # 为重复字段生成带编号的标签
    label_counts_batch = {}
    for field in fields:
        lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
        label_counts_batch[lbl] = label_counts_batch.get(lbl, 0) + 1

    label_seen_batch = {}
    field_map = {}
    for field in fields:
        lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
        if label_counts_batch[lbl] > 1:
            label_seen_batch[lbl] = label_seen_batch.get(lbl, 0) + 1
            display_lbl = f"{lbl} #{label_seen_batch[lbl]}"
        else:
            display_lbl = lbl
        if display_lbl in headers:
            field_map[display_lbl] = {'col': headers[display_lbl], 'field': field}
        elif lbl in headers:
            # 向后兼容：没编号的旧模板也能匹配
            field_map[lbl] = {'col': headers[lbl], 'field': field}

    if not field_map:
        raise HTTPException(status_code=400, detail="Excel表头与模板字段不匹配，请使用系统生成的模板")

    # 读取数据行（第4行起）
    data_rows = []
    for row_idx in range(4, ws.max_row + 1):
        row_data = {}
        has_data = False
        for label, info in field_map.items():
            cell_val = ws.cell(row=row_idx, column=info['col']).value
            if cell_val is not None and str(cell_val).strip():
                row_data[label] = str(cell_val).strip()
                has_data = True
            else:
                row_data[label] = ''
        if has_data:
            data_rows.append(row_data)

    if not data_rows:
        raise HTTPException(status_code=400, detail="Excel中没有填写数据（请从第4行开始填写）")

    if len(data_rows) > 100:
        raise HTTPException(status_code=400, detail="单次批量上传最多支持100行数据")

    # 获取PDF原始数据（优先文件系统）
    pdf_bytes_orig = None
    if tmpl.get("pdf_path"):
        try:
            pdf_bytes_orig = _read_pdf_file(tmpl["pdf_path"])
        except FileNotFoundError:
            pass
    if pdf_bytes_orig is None and tmpl.get("pdf_data"):
        pdf_bytes_orig = bytes(tmpl["pdf_data"]) if not isinstance(tmpl["pdf_data"], bytes) else tmpl["pdf_data"]
    if not pdf_bytes_orig:
        raise HTTPException(status_code=500, detail="PDF模板文件不存在")

    # 批量生成PDF
    try:
        import pypdf
        from pypdf import PdfWriter, PdfReader
        HAS_PYPDF_LOCAL = True
    except ImportError:
        HAS_PYPDF_LOCAL = False

    zip_buffer = io.BytesIO()
    generated_count = 0
    errors = []

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for row_idx, row_data in enumerate(data_rows):
            try:
                # 使用reportlab在PDF上叠加文字
                filled_pdf = _fill_pdf_with_data(pdf_bytes_orig, fields, row_data)
                holder_name = _extract_holder_name(row_data)
                seq_num = row_idx + 1
                filename = f"{holder_name}_{tmpl['name']}_{seq_num:03d}.pdf" if holder_name else f"{tmpl['name']}_{seq_num:03d}.pdf"
                zf.writestr(filename, filled_pdf)
                generated_count += 1
            except Exception as e:
                errors.append(f"第{row_idx + 4}行：{str(e)}")

    zip_buffer.seek(0)

    # 返回结果
    if generated_count == 0:
        raise HTTPException(status_code=500, detail=f"所有PDF生成失败：{'; '.join(errors[:3])}")

    zip_filename = f"{tmpl['name']}_批量PDF_{generated_count}份.zip"
    encoded_zip_filename = quote(zip_filename)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_zip_filename}",
            "X-Generated-Count": str(generated_count),
            "X-Error-Count": str(len(errors)),
        }
    )


def _extract_holder_name(row_data: dict) -> str:
    """从填写数据中提取保单持有人名（尝试匹配常见的姓名字段标签）"""
    # 按优先级匹配可能的姓名字段关键词
    name_keywords = [
        '保单持有人', '持有人', '投保人', '被保险人', '保险人',
        '持有人姓名', '投保人姓名', '被保人',
        '姓名', '名字', '全名', '客户姓名', '客户名',
        'Name', 'name', 'Full Name', 'full_name',
        'Policyholder', 'policyholder',
    ]
    # 精确匹配
    for keyword in name_keywords:
        for label, value in row_data.items():
            if label.strip() == keyword and value and str(value).strip():
                return str(value).strip()
    # 模糊匹配（标签中包含关键词）
    for keyword in name_keywords:
        for label, value in row_data.items():
            if keyword in label and value and str(value).strip():
                return str(value).strip()
    return ''


def _download_cjk_font() -> str:
    """下载完整CJK字体到临时目录并返回路径（支持简繁体中文）"""
    import tempfile
    import urllib.request
    
    font_cache_path = os.path.join(tempfile.gettempdir(), 'NotoSansCJKtc-Regular.otf')
    if os.path.exists(font_cache_path) and os.path.getsize(font_cache_path) > 500000:
        return font_cache_path
    
    # 使用完整的CJK字体（非子集），TC优先（繁体，同时包含简体字形）
    font_urls = [
        'https://cdn.jsdelivr.net/gh/googlefonts/noto-cjk@main/Sans/OTF/TraditionalChinese/NotoSansCJKtc-Regular.otf',
        'https://cdn.jsdelivr.net/gh/googlefonts/noto-cjk@main/Sans/OTF/SimplifiedChinese/NotoSansCJKsc-Regular.otf',
        'https://raw.githubusercontent.com/googlefonts/noto-cjk/main/Sans/OTF/TraditionalChinese/NotoSansCJKtc-Regular.otf',
    ]
    
    for url in font_urls:
        try:
            print(f"正在下载完整CJK字体: {url}")
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=60) as resp:
                font_data = resp.read()
                # 完整CJK字体应该大于1MB，子集字体通常只有几十KB
                if len(font_data) > 500000:
                    with open(font_cache_path, 'wb') as f:
                        f.write(font_data)
                    print(f"完整CJK字体下载成功: {font_cache_path} ({len(font_data)} bytes)")
                    return font_cache_path
                else:
                    print(f"字体文件过小（可能是子集），跳过: {url} ({len(font_data)} bytes)")
        except Exception as e:
            print(f"字体下载失败 {url}: {e}")
            continue
    
    return None


# 全局缓存：CJK字体路径
_cached_cjk_font_path = None


def _get_cjk_font_path() -> str:
    """获取完整CJK字体路径（优先本地，备选下载）"""
    global _cached_cjk_font_path
    if _cached_cjk_font_path and os.path.exists(_cached_cjk_font_path):
        # 验证缓存字体文件大小（完整CJK字体应大于500KB）
        if os.path.getsize(_cached_cjk_font_path) > 500000:
            return _cached_cjk_font_path
        else:
            _cached_cjk_font_path = None
    
    # 尝试本地字体（TTF/OTF 单体优先，TTC 放后面因为 reportlab 不一定支持）
    local_font_paths = [
        # TTF 单体字体（reportlab 完美支持）
        '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttf',
        '/usr/share/fonts/truetype/wqy/wqy-microhei.ttf',
        '/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf',
        '/usr/share/fonts/opentype/noto/NotoSansCJKtc-Regular.otf',
        '/usr/share/fonts/truetype/noto/NotoSansCJKsc-Regular.otf',
        '/usr/share/fonts/truetype/noto/NotoSansCJKtc-Regular.otf',
        # TTC 集合字体（reportlab 可能不支持 PostScript outlines）
        '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
        '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/opentype/noto-cjk/NotoSansCJK-Regular.ttc',
        # macOS系统字体
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/Library/Fonts/Arial Unicode.ttf',
    ]
    for fp in local_font_paths:
        if os.path.exists(fp):
            # 验证字体文件大小
            try:
                if os.path.getsize(fp) > 100000:  # 至少100KB
                    _cached_cjk_font_path = fp
                    print(f"找到本地中文字体: {fp} ({os.path.getsize(fp)} bytes)")
                    return fp
            except:
                continue
    
    # 本地无字体，从CDN下载
    downloaded = _download_cjk_font()
    if downloaded:
        _cached_cjk_font_path = downloaded
        return downloaded
    
    return None


def _has_cjk(text):
    """检测文本是否包含CJK字符"""
    return bool(re.search(r'[\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff\u2e80-\u2eff\u3000-\u303f\ufe30-\ufe4f]', text))


def _draw_mixed_text(c, x, y, text, cjk_font_name, font_size, font_registered):
    """混合渲染：CJK字符用CJK字体，ASCII字符用Helvetica
    
    解决CJK字体中数字/字母显示过宽的问题（全角宽度）
    同时确保文字渲染模式为仅填充（mode=0），防止黑框
    """
    cursor_x = x
    # 按连续的CJK / 非CJK字符分段
    segments = re.findall(
        r'([\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff\u2e80-\u2eff\u3000-\u303f\ufe30-\ufe4f]+|'
        r'[^\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff\u2e80-\u2eff\u3000-\u303f\ufe30-\ufe4f]+)',
        text
    )
    
    for seg in segments:
        if _has_cjk(seg) and font_registered:
            c.setFont(cjk_font_name, font_size)
        else:
            c.setFont('Helvetica', font_size)
        
        # 确保每段渲染前都设置正确的颜色和渲染模式，防止黑框
        c.setFillColorRGB(0, 0, 0)       # 填充色：黑色
        c.setStrokeColorRGB(0, 0, 0)     # 描边色也设黑色（以防万一）
        c.setLineWidth(0)                 # 线宽为0
        # 强制文字渲染模式为 0（仅填充，不描边）
        c._textRenderMode = 0
        
        c.drawString(cursor_x, y, seg)
        # 计算当前段宽度，移动光标
        seg_width = c.stringWidth(seg, c._fontname, font_size)
        cursor_x += seg_width


def _fill_pdf_with_data(pdf_bytes: bytes, fields: list, row_data: dict) -> bytes:
    """将数据填入PDF并返回填写后的PDF字节"""
    try:
        from pypdf import PdfWriter, PdfReader
        import pypdf
    except ImportError:
        raise Exception("pypdf not available")

    # 尝试使用reportlab叠加文字
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from pypdf import PdfWriter, PdfReader

        reader = PdfReader(io.BytesIO(pdf_bytes))
        writer = PdfWriter()

        # 预先注册中文字体（在循环外只注册一次）
        font_registered = False
        cjk_font_name = 'CJKFont'
        try:
            cjk_font_path = _get_cjk_font_path()
            if cjk_font_path:
                try:
                    # 检查是否已注册（避免重复注册报错）
                    try:
                        pdfmetrics.getFont(cjk_font_name)
                        font_registered = True
                    except KeyError:
                        pdfmetrics.registerFont(TTFont(cjk_font_name, cjk_font_path))
                        font_registered = True
                    print(f"CJK字体注册成功: {cjk_font_path}")
                except Exception as reg_err:
                    print(f"CJK字体注册失败: {reg_err}")
        except Exception as e:
            print(f"获取CJK字体路径异常: {e}")

        # 构建字段值映射（支持带编号的标签和普通标签）
        # 将 row_data 中的键标准化：去掉 " #N" 后缀用于匹配
        field_label_counts = {}
        for field in fields:
            lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
            field_label_counts[lbl] = field_label_counts.get(lbl, 0) + 1

        field_label_seen = {}
        field_value_map = {}
        for field in fields:
            lbl = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
            if field_label_counts[lbl] > 1:
                field_label_seen[lbl] = field_label_seen.get(lbl, 0) + 1
                display_lbl = f"{lbl} #{field_label_seen[lbl]}"
            else:
                display_lbl = lbl
            # 优先精确匹配，否则尝试去掉编号后匹配
            if display_lbl in row_data:
                field_value_map[display_lbl] = row_data[display_lbl]
            elif lbl in row_data:
                field_value_map[display_lbl] = row_data[lbl]
            else:
                field_value_map[display_lbl] = ''

        # 按页分组字段
        pages_fields = {}
        for field in fields:
            label = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
            page_num = field.get('page_num', 1)
            if page_num not in pages_fields:
                pages_fields[page_num] = []
            pages_fields[page_num].append((label, field))

        # 为每个字段分配索引（用于处理重复字段的编号）
        field_indices = {}
        field_index_counts = {}
        for idx, field in enumerate(fields):
            lbl_raw = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
            if lbl_raw not in field_index_counts:
                field_index_counts[lbl_raw] = 0
            field_index_counts[lbl_raw] += 1
            field_indices[idx] = field_index_counts[lbl_raw]

        for page_idx in range(len(reader.pages)):
            page_num = page_idx + 1
            orig_page = reader.pages[page_idx]
            page_width = float(orig_page.mediabox.width)
            page_height = float(orig_page.mediabox.height)

            # 创建叠加层
            overlay_buffer = io.BytesIO()
            c = rl_canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))

            if page_num in pages_fields:
                for label, field in pages_fields[page_num]:
                    # 获取field在fields列表中的索引
                    try:
                        field_global_idx = fields.index(field)
                    except ValueError:
                        continue
                    
                    # 获取display_lbl编号
                    lbl_raw = field.get('custom_label') or field.get('field_label') or field.get('field_name') or ''
                    total_count = field_label_counts.get(lbl_raw, 0)
                    field_seq_num = field_indices.get(field_global_idx, 1)
                    
                    if total_count > 1:
                        display_lbl = f"{lbl_raw} #{field_seq_num}"
                    else:
                        display_lbl = lbl_raw
                    
                    # 从field_value_map获取值
                    value = field_value_map.get(display_lbl, '')
                    if not value:
                        continue

                    x = float(field.get('x', 0))
                    y = float(field.get('y', 0))
                    w = float(field.get('width', 100))
                    h = float(field.get('height', 20))
                    font_size = int(field.get('font_size', 12))
                    field_type = field.get('field_type', 'text')

                    # 处理复选框
                    if field_type == 'checkbox':
                        display_val = '☑' if value in ('是', 'true', 'True', '1', 'yes', 'Yes') else '☐'
                    else:
                        display_val = str(value)

                    # 前端 screenToPdf 已经将坐标转为 PDF 坐标系（原点左下角）
                    # reportlab 也是左下角坐标系，所以直接使用 y 即可
                    # 垂直居中：在框内偏移使文字居中
                    text_y = y + (h - font_size) / 2 + 2
                    
                    # 确保文字在页面范围内
                    if text_y < 0:
                        text_y = 5
                    if x < 0:
                        x = 2

                    # 设置填充色为黑色，禁用描边防止黑框
                    c.setFillColorRGB(0, 0, 0)
                    c.setStrokeColorRGB(0, 0, 0)
                    c.setLineWidth(0)
                    # 强制文字渲染模式为仅填充（mode=0），防止黑框
                    c._textRenderMode = 0
                    
                    try:
                        # 使用混合渲染：CJK字符用CJK字体，ASCII字符用Helvetica
                        # 解决CJK字体中数字/字母显示过宽的问题
                        _draw_mixed_text(c, x + 2, text_y, display_val, cjk_font_name, font_size, font_registered)
                    except Exception as draw_err:
                        print(f"绘制文字失败: {display_val}, 错误: {draw_err}")
                        # 降级：直接用Helvetica绘制
                        try:
                            c.setFont('Helvetica', font_size)
                            c.setFillColorRGB(0, 0, 0)
                            c._textRenderMode = 0
                            safe_text = str(display_val)[:200]
                            c.drawString(x + 2, text_y, safe_text)
                        except Exception as safe_err:
                            print(f"安全绘制也失败: {safe_err}")

            c.save()
            overlay_buffer.seek(0)

            # 合并叠加层
            overlay_reader = PdfReader(overlay_buffer)
            if overlay_reader.pages:
                orig_page.merge_page(overlay_reader.pages[0])
            writer.add_page(orig_page)

        output = io.BytesIO()
        writer.write(output)
        return output.getvalue()

    except Exception as e:
        # 降级：直接返回原始PDF（不填写内容）
        raise Exception(f"PDF填写失败: {str(e)}")


@app.get("/api/templates/{tid}/batch-preview")
def batch_preview_info(tid: int):
    """获取批量上传预览信息（字段列表）"""
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, name FROM templates WHERE id=%s", (tid,))
            tmpl = cursor.fetchone()
            if not tmpl:
                raise HTTPException(status_code=404, detail="Template not found")
            cursor.execute("""
                SELECT tf.id, tf.page_num, tf.font_size,
                       COALESCE(tf.custom_label, fd.label, fd.name, '未命名') as label,
                       COALESCE(fd.field_type, 'text') as field_type
                FROM template_fields tf
                LEFT JOIN field_definitions fd ON tf.field_def_id = fd.id
                WHERE tf.template_id = %s
                ORDER BY tf.page_num, tf.sort_order, tf.id
            """, (tid,))
            fields = cursor.fetchall()
    finally:
        db.close()
    return {"data": {"template": tmpl, "fields": fields, "field_count": len(fields)}}


# ==================== COS存储管理API ====================

@app.get("/api/storage/status")
def storage_status():
    """查看PDF存储状态：有多少在COS中，多少在数据库中"""
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) as total FROM templates")
            total = cursor.fetchone()["total"]
            cursor.execute("SELECT COUNT(*) as cnt FROM templates WHERE pdf_path IS NOT NULL AND pdf_path != ''")
            cos_count = cursor.fetchone()["cnt"]
            cursor.execute("SELECT COUNT(*) as cnt FROM templates WHERE pdf_data IS NOT NULL AND (pdf_path IS NULL OR pdf_path = '')")
            db_count = cursor.fetchone()["cnt"]
    finally:
        db.close()

    # 统计COS中的文件数
    cos_objects_count = 0
    cos_size_total = 0
    try:
        client = _get_cos_client()
        marker = ""
        while True:
            response = client.list_objects(
                Bucket=COS_BUCKET,
                Prefix=COS_PREFIX,
                Marker=marker,
                MaxKeys=1000,
            )
            contents = response.get("Contents", [])
            for obj in contents:
                cos_objects_count += 1
                cos_size_total += int(obj.get("Size", 0))
            if response.get("IsTruncated") == "true":
                marker = response.get("NextMarker", "")
            else:
                break
    except Exception as e:
        print(f"列举COS对象失败: {e}")

    return {
        "data": {
            "total_templates": total,
            "stored_in_cos": cos_count,
            "stored_in_database": db_count,
            "cos_bucket": COS_BUCKET,
            "cos_region": COS_REGION,
            "cos_prefix": COS_PREFIX,
            "cos_objects_count": cos_objects_count,
            "cos_usage_mb": round(cos_size_total / (1024 * 1024), 2),
        }
    }


@app.post("/api/storage/migrate")
def trigger_migration():
    """手动触发将数据库中的PDF数据迁移到COS"""
    if not COS_SECRET_ID or not COS_SECRET_KEY or not COS_BUCKET:
        raise HTTPException(status_code=400, detail="COS未配置，请先设置 COS_SECRET_ID / COS_SECRET_KEY / COS_BUCKET 环境变量")
    db = get_db()
    try:
        _migrate_pdf_to_cos(db)
    finally:
        db.close()
    return {"success": True, "message": "迁移到COS完成"}


@app.get("/")
def root():
    return RedirectResponse(url="/static/index.html")


# ==================== COS模板导入API ====================

@app.get("/api/cos/list")
def list_cos_files(prefix: str = ""):
    """列举COS中的PDF文件（用于导入已有模板）"""
    client = _get_cos_client()
    search_prefix = prefix if prefix else ""

    objects = []
    marker = ""
    while True:
        response = client.list_objects(
            Bucket=COS_BUCKET,
            Prefix=search_prefix,
            Marker=marker,
            MaxKeys=200,
        )
        contents = response.get("Contents", [])
        for obj in contents:
            key = obj.get("Key", "")
            # 只列出PDF文件
            if key.lower().endswith(".pdf"):
                objects.append({
                    "key": key,
                    "name": key.split("/")[-1],
                    "size": int(obj.get("Size", 0)),
                    "last_modified": obj.get("LastModified", ""),
                })
        if response.get("IsTruncated") == "true":
            marker = response.get("NextMarker", "")
        else:
            break

    return {"data": objects, "bucket": COS_BUCKET, "region": COS_REGION}


@app.post("/api/cos/import")
def import_cos_template(body: dict):
    """将COS中已有的PDF文件导入为模板"""
    cos_key = body.get("cos_key", "").strip()
    name = body.get("name", "").strip()
    description = body.get("description", "").strip()

    if not cos_key:
        raise HTTPException(status_code=400, detail="cos_key不能为空")
    if not name:
        # 从文件名自动生成模板名称
        name = cos_key.split("/")[-1].replace(".pdf", "").replace("_", " ")

    # 验证文件存在并可读取
    client = _get_cos_client()
    try:
        head = client.head_object(Bucket=COS_BUCKET, Key=cos_key)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"COS文件不存在: {cos_key}, 错误: {e}")

    pdf_filename = cos_key.split("/")[-1]

    # 直接将COS Key写入数据库（不重新上传，复用现有文件）
    db = get_db()
    try:
        with db.cursor() as cursor:
            cursor.execute(
                "INSERT INTO templates (name, description, pdf_path, pdf_filename, page_count) VALUES (%s, %s, %s, %s, %s)",
                (name, description, cos_key, pdf_filename, 1)
            )
            db.commit()
            new_id = cursor.lastrowid
            cursor.execute(
                "SELECT id, name, description, pdf_filename, pdf_path, page_count, created_at FROM templates WHERE id=%s",
                (new_id,)
            )
            row = cursor.fetchone()
        return {"data": row, "message": f"已成功导入COS模板: {cos_key}"}
    finally:
        db.close()


app.mount("/static", StaticFiles(directory="static", html=True), name="static")
